# 프로세스 설명:
# 0. 엑셀 데이터 가져오기
#    - 엑셀 파일(processed_transactions.xlsx)에서 C열(이름) 데이터 읽기

# 1. 데이터 가져오기 (get_data_from_sheets 함수)
#    - Google 인증 정보를 가져와 Sheets API에 연결
#    - 지정된 스프레드시트의 모든 시트를 순회
#    - 각 시트에서 F8:J300 범위의 데이터를 가져옴
#    - 각 행에서 F열(이름), I열(계좌번호), J열(주민번호) 데이터 수집
#    - 이름을 키로, (계좌번호, 주민번호)를 값으로 하는 딕셔너리 생성
# 
# 2. 엑셀 파일 업데이트 (update_excel 함수)
#    - 지정된 엑셀 파일을 열고 활성 시트에 접근
#    - 각 행을 순회하면서 C열에서 이름을 가져옴
#    - 스프레드시트에서 가져온 데이터에 해당 이름이 있으면 F열에 계좌번호, G열에 주민번호 입력
#    - 이름이 없으면 두 열에 'NAN' 입력
#    - 변경사항 저장
# 
# 3. 메인 실행 (main 함수)
#    - 스프레드시트 ID와 처리할 엑셀 파일명 지정
#    - 로컬 엑셀에서 이름 데이터를 가져옴
#    - 스프레드시트에서 해당 이름의 데이터를 가져옴
#    - 데이터 가져오기에 성공하면 엑셀 파일을 업데이트
#    - 실패하면 오류 메시지를 기록하고 종료
# 모든 과정에서 로깅을 통해 진행 상황과 오류를 기록

import openpyxl
from auth import get_credentials
from googleapiclient.discovery import build
import logging
import datetime
import re
from collections import defaultdict

# 일반 로깅 설정
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# 중복 케이스를 위한 별도 로거 설정
# duplicate_logger = logging.getLogger('duplicate_logger')
# duplicate_logger.setLevel(logging.WARNING)

# 매번 새 파일을 생성하는 대신 고정된 파일명 사용
# duplicate_log_file = 'duplicate_cases.log'

# 파일 핸들러 설정 - 'a' 모드로 파일 열기 (추가 모드)
# file_handler = logging.FileHandler(duplicate_log_file, mode='a', encoding='utf-8')
# file_handler.setFormatter(logging.Formatter('%(asctime)s - %(levelname)s - %(message)s'))
# duplicate_logger.addHandler(file_handler)

def read_excel_data(excel_file):
    """로컬 엑셀 파일에서 C열(이름)과 E열(제품) 데이터를 함께 읽어옵니다."""
    logging.info(f"엑셀 파일 '{excel_file}'에서 이름과 제품 데이터 읽기 시작")
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active
        
        names = []
        name_product_map = {}  # 이름과 제품을 매핑하는 딕셔너리
        
        for row in range(2, sheet.max_row + 1):  # 첫 번째 행은 헤더이므로 두 번째 행부터 시작
            name = sheet.cell(row=row, column=3).value  # C열 (이름)
            product = sheet.cell(row=row, column=5).value  # E열 (제품)
            
            if name:
                names.append(name)
                name_product_map[name] = product or "정보 없음"
        
        logging.info(f"엑셀 파일에서 {len(names)}개의 이름과 제품 데이터 읽기 완료")
        return names, name_product_map
    except Exception as e:
        logging.error(f"엑셀 파일에서 데이터 읽기 오류: {str(e)}")
        return [], {}

def get_data_from_sheets(spreadsheet_id, names_to_find):
    logging.info("스프레드시트에서 데이터 가져오기 시작")
    try:
        creds = get_credentials()
        service = build('sheets', 'v4', credentials=creds)

        # 스프레드시트의 모든 시트 가져오기
        sheet_metadata = service.spreadsheets().get(spreadsheetId=spreadsheet_id).execute()
        sheets = sheet_metadata.get('sheets', '')

        all_data = {}  # 기본 데이터 (엑셀 업데이트용)
        
        # 동명이인 정보 수집 구조: {이름: {주민번호: {시트1, 시트2, ...}}}
        name_id_sheet_map = defaultdict(lambda: defaultdict(set))
        
        for sheet in sheets:
            sheet_name = sheet['properties']['title']
            range_name = f"'{sheet_name}'!F8:J300"
            result = service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id, range=range_name).execute()
            values = result.get('values', [])
            
            for row in values:
                if len(row) >= 5:
                    name = row[0]  # F열 (이름)
                    # 우리가 찾는 이름 목록에 있는 경우만 처리
                    if name in names_to_find:
                        account = row[3] if len(row) > 3 else 'NAN'  # I열 (계좌번호)
                        id_number = row[4] if len(row) > 4 else 'NAN'  # J열 (주민번호)
                        
                        # 주민번호 형식 간단 검증 (6자리-7자리 형식)
                        is_valid_id = id_number != 'NAN' and '-' in id_number and len(id_number.replace('-', '')) >= 12
                        
                        if is_valid_id:
                            # 동명이인 정보 수집 - 각 이름/주민번호 조합에 대해 발견된 시트 정보 저장
                            name_id_sheet_map[name][id_number].add(sheet_name)
                        
                        # 처음 발견된 유효한 데이터를 기본 값으로 사용
                        if name not in all_data and is_valid_id:
                            all_data[name] = (account, id_number)
                        # 이미 발견된 이름이지만 주민번호가 다른 경우는 중복 표시
                        elif name in all_data and is_valid_id:
                            _, existing_id = all_data[name]
                            if existing_id != id_number and existing_id != 'DUPLICATE':
                                duplicate_msg = f"동명이인 발견: '{name}' - 주민번호 불일치 (기존: {existing_id}, 새로운 값: {id_number}, 시트: {sheet_name})"
                                logging.warning(duplicate_msg)
                                # duplicate_logger.warning(duplicate_msg)  # 이 줄을 제거했습니다.
                                all_data[name] = ('DUPLICATE', 'DUPLICATE')

        # 동명이인 정보 파일로 저장
        save_name_id_mapping(name_id_sheet_map)
        
        logging.info("스프레드시트에서 데이터 가져오기 완료")
        return all_data
    except Exception as e:
        logging.error(f"스프레드시트 데이터 가져오기 오류: {str(e)}")
        return None

def save_name_id_mapping(name_id_sheet_map):
    """동명이인 분석을 위한 매핑 정보를 파일로 저장합니다."""
    try:
        import json
        
        # JSON 직렬화를 위해 defaultdict를 일반 dict로 변환
        serializable_map = {}
        for name, id_map in name_id_sheet_map.items():
            serializable_map[name] = {id_num: list(sheets) for id_num, sheets in id_map.items()}
        
        with open('name_id_mapping.json', 'w', encoding='utf-8') as f:
            json.dump(serializable_map, f, ensure_ascii=False, indent=2)
        
        logging.info("이름-주민번호-시트 매핑 정보가 저장되었습니다.")
    except Exception as e:
        logging.error(f"매핑 정보 저장 중 오류 발생: {str(e)}")

def update_excel(excel_file, sheet_data):
    logging.info(f"엑셀 파일 '{excel_file}' 업데이트 시작")
    try:
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in range(2, sheet.max_row + 1):  # 첫 번째 행은 헤더이므로 두 번째 행부터 시작
            name = sheet.cell(row=row, column=3).value  # C열 (이름)
            if name in sheet_data:
                account, id_number = sheet_data[name]
                sheet.cell(row=row, column=6, value=account)  # F열 (계좌번호)
                sheet.cell(row=row, column=7, value=id_number)  # G열 (주민번호)
            else:
                sheet.cell(row=row, column=6, value='NAN')
                sheet.cell(row=row, column=7, value='NAN')

        wb.save(excel_file)
        logging.info(f"엑셀 파일 '{excel_file}' 업데이트 완료")
    except Exception as e:
        logging.error(f"엑셀 파일 업데이트 오류: {str(e)}")

def print_duplicate_summary(name_product_map):
    """동명이인 정보를 사용자 친화적으로 출력하고 파일에 저장합니다."""
    try:
        import json
        
        # 저장된 매핑 정보 읽기
        with open('name_id_mapping.json', 'r', encoding='utf-8') as f:
            name_id_sheet_map = json.load(f)
        
        summary_lines = ["\n===== 동명이인 요약 보고서 =====\n"]
        
        # 동명이인 문제로 처리된 이름 목록
        duplicate_names = []
        
        for name, id_map in name_id_sheet_map.items():
            # 주민번호가 2개 이상일 때만 동명이인으로 처리
            # 주민번호 표준화 (공백 제거)
            standardized_id_map = {}
            for id_num, sheets in id_map.items():
                # 주민번호에서 공백 제거하고 표준화
                clean_id = id_num.strip()
                if clean_id in standardized_id_map:
                    # 이미 있는 주민번호면 시트 정보 병합
                    for sheet in sheets:
                        standardized_id_map[clean_id].append(sheet)
                else:
                    standardized_id_map[clean_id] = sheets
            
            # 표준화된 주민번호 맵으로 동명이인 처리
            if len(standardized_id_map) >= 2:
                duplicate_names.append(name)
                
                # 엑셀 정보 표시
                summary_lines.append(f"[엑셀(C열)]")
                summary_lines.append(f"{name}")
                product = name_product_map.get(name, "정보 없음")
                summary_lines.append(f"제품 : [{product}]\n")
                
                # 각 후보(주민번호) 정보 표시
                for i, (id_num, sheets) in enumerate(standardized_id_map.items()):
                    # 중복 시트 제거하고 정렬
                    unique_sheets = sorted(set(sheets))
                    sheets_str = ", ".join(unique_sheets)
                    
                    summary_lines.append(f"후보{i+1}")
                    summary_lines.append(f"{name}")
                    summary_lines.append(f"{id_num}")
                    summary_lines.append(f"시트: [{sheets_str}]\n")
                
                summary_lines.append("-" * 40)
        
        # 총 동명이인 개수 추가
        total_duplicates = len(duplicate_names)
        summary_lines.append(f"\n총 {total_duplicates}명의 동명이인이 발견되었습니다.")
        if total_duplicates > 0:
            duplicate_names_str = ", ".join(duplicate_names)
            summary_lines.append(f"확인이 필요한 이름: {duplicate_names_str}")
        summary_lines.append("\n")
        
        # 요약 보고서를 터미널에 출력
        for line in summary_lines:
            print(line)
        
        # 요약 보고서를 파일에 저장 (덮어쓰기 모드)
        summary_file = 'duplicate_summary.txt'  # 고정된 파일 이름
        with open(summary_file, 'w', encoding='utf-8') as f:
            for line in summary_lines:
                f.write(line + '\n')
        logging.info(f"동명이인 요약 보고서가 '{summary_file}' 파일에 저장되었습니다.")
        
        # 최종 요약 로그 추가
        logging.info(f"총 {total_duplicates}명의 동명이인 발견 - 확인 필요")
    except Exception as e:
        logging.error(f"요약 보고서 생성 중 오류 발생: {str(e)}")

def main():
    spreadsheet_id = '1CK2UXTy7HKjBe2T0ovm5hfzAAKZxZAR_ev3cbTPOMPs'
    excel_file = 'processed_transactions.xlsx'  # 실제 엑셀 파일 이름으로 변경해주세요

    # 1. 로컬 엑셀에서 이름과 제품 목록 가져오기
    names, name_product_map = read_excel_data(excel_file)
    if not names:
        logging.error("엑셀 파일에서 이름 데이터를 가져오지 못했습니다. 프로그램을 종료합니다.")
        return

    # 2. 스프레드시트에서 해당 이름에 대한 데이터 가져오기
    sheet_data = get_data_from_sheets(spreadsheet_id, names)
    if sheet_data:
        # 3. 로컬 엑셀 파일 업데이트
        update_excel(excel_file, sheet_data)
        
        # 4. 중복 케이스 요약 보고서 생성
        logging.info("동명이인 요약 보고서 생성 중...")
        print_duplicate_summary(name_product_map)
    else:
        logging.error("스프레드시트 데이터를 가져오지 못했습니다. 프로그램을 종료합니다.")

if __name__ == "__main__":
    main()