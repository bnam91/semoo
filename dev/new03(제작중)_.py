import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# 엑셀 파일 경로
file_path = r"C:\Users\USER\Desktop\github\semoo\(전처리)구글시트_2508_20250909_112248.xlsx"

try:
    # 후처리 시트 데이터 읽기
    df = pd.read_excel(file_path, sheet_name='후처리')
    
    # 새로운 데이터프레임 생성
    new_df = pd.DataFrame()
    
    # 열 순서 변경
    new_df[2] = df.iloc[:, 7]  # H열 -> 인덱스 2 (날짜)
    new_df[3] = df.iloc[:, 1]  # B열 -> 인덱스 3 (텍스트)
    new_df[4] = df.iloc[:, 7]  # H열 -> 인덱스 4 (숫자)
    new_df[9] = df.iloc[:, 6]  # G열 -> 인덱스 9 (숫자)
    
    # 신규 데이터 입력
    new_df[7] = '기타자영업'  # 인덱스 7
    new_df[10] = 0.033  # 인덱스 10
    
    # 엑셀 파일 로드
    wb = load_workbook(file_path)
    
    # '전달준비' 시트가 없으면 생성
    if '전달준비' not in wb.sheetnames:
        wb.create_sheet('전달준비')
    
    # 데이터프레임을 엑셀에 저장
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        new_df.to_excel(writer, sheet_name='전달준비', index=False, header=False)
    
    # 엑셀 파일 다시 로드
    wb = load_workbook(file_path)
    ws = wb['전달준비']
    
    # 함수식 추가
    for row in range(2, ws.max_row + 1):  # 헤더 제외하고 시작
        # 인덱스 11의 함수식 (인덱스 9 * 인덱스 10)
        ws.cell(row=row, column=12).value = f'={get_column_letter(10)}{row}*{get_column_letter(11)}{row}'
        
        # 인덱스 12의 함수식 (인덱스 11 * 0.1)
        ws.cell(row=row, column=13).value = f'={get_column_letter(12)}{row}*0.1'
    
    # 변경사항 저장
    wb.save(file_path)
    print("작업이 완료되었습니다.")

except Exception as e:
    print("오류 발생:", str(e))
