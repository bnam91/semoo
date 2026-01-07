#!/usr/bin/env python3
"""
가상환경 자동 감지 및 활성화
"""
import sys
import os
from pathlib import Path

# 현재 스크립트의 디렉토리
SCRIPT_DIR = Path(__file__).parent.absolute()
VENV_PYTHON = SCRIPT_DIR / 'venv' / 'bin' / 'python3'

# 가상환경이 존재하고 현재 Python이 가상환경이 아닌 경우
if VENV_PYTHON.exists() and 'venv' not in sys.executable:
    # 가상환경의 Python으로 재실행
    os.execv(str(VENV_PYTHON), [str(VENV_PYTHON)] + sys.argv)

import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import os

# 현재 폴더의 xlsx 파일 목록 가져오기
xlsx_파일들 = [f for f in os.listdir('.') if f.endswith('.xlsx')]

if not xlsx_파일들:
    print("현재 폴더에 xlsx 파일이 없습니다.")
    exit()

# 파일 목록 출력
print("\n사용 가능한 엑셀 파일 목록:")
for i, 파일명 in enumerate(xlsx_파일들, 1):
    print(f"{i}. {파일명}")

# 사용자 입력 받기
while True:
    try:
        선택 = int(input("\n처리할 파일 번호를 입력하세요: "))
        if 1 <= 선택 <= len(xlsx_파일들):
            파일경로 = xlsx_파일들[선택-1]
            break
        else:
            print("올바른 번호를 입력해주세요.")
    except ValueError:
        print("숫자를 입력해주세요.")

# 엑셀 파일 열기
워크북 = openpyxl.load_workbook(파일경로)

# 원본 시트 가져오기
원본시트 = 워크북['25년3월데이터']

# 새 시트 생성 (이미 있으면 삭제 후 생성)
if '후처리' in 워크북.sheetnames:
    워크북.remove(워크북['후처리'])
후처리시트 = 워크북.create_sheet('후처리')

# 동명이인 시트 생성 (이미 있으면 삭제 후 생성)
if '동명이인' in 워크북.sheetnames:
    워크북.remove(워크북['동명이인'])
동명이인시트 = 워크북.create_sheet('동명이인')

# 오렌지색 채우기 스타일 정의
오렌지색 = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
# 파란색 채우기 스타일 정의
파란색 = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

# 헤더 행 복사
헤더행 = next(원본시트.rows)
for 셀 in 헤더행:
    후처리시트.cell(row=1, column=셀.column, value=셀.value)
    동명이인시트.cell(row=1, column=셀.column, value=셀.value)

# 동일인 데이터 그룹화 (이름과 주민번호가 같은 경우)
사람별_데이터 = defaultdict(lambda: {'A열값들': [], 'G열합계': 0, '행데이터들': [], 'H열값들': []})

# 이름에서 괄호 처리 함수
def 이름_정규화(이름):
    """이름에서 괄호를 제거하고 괄호 안의 이름도 고려하여 정규화"""
    if not 이름 or not isinstance(이름, str):
        return ""
    
    # 전체 이름에서 공백 제거
    이름 = 이름.replace(' ', '')
    
    # 괄호가 있는 경우
    if '(' in 이름 and ')' in 이름:
        # 괄호 밖 이름과 괄호 안 이름 모두 추출
        괄호밖_이름 = 이름.split('(')[0].strip()
        괄호안_이름 = 이름.split('(')[1].split(')')[0].strip()
        return [괄호밖_이름, 괄호안_이름]
    else:
        # 괄호가 없는 경우
        return [이름.strip()]

# 첫 번째 행은 이미 처리했으므로 두 번째 행부터 시작
for 행 in list(원본시트.rows)[1:]:
    # B열(이름)과 F열(주민번호) 값 가져오기
    원본_이름 = 행[1].value if len(행) > 1 else ""
    주민번호 = 행[5].value if len(행) > 5 else ""
    
    # 이름 정규화
    정규화된_이름들 = 이름_정규화(원본_이름)
    
    # A열 값과 G열 값, H열 값 가져오기
    A열값 = 행[0].value if len(행) > 0 else ""
    G열값 = 행[6].value if len(행) > 6 else 0
    H열값 = 행[7].value if len(행) > 7 else ""
    
    # 숫자가 아닌 경우 처리 - 쉼표 제거 및 숫자 변환 개선
    if not isinstance(G열값, (int, float)):
        try:
            if isinstance(G열값, str):
                # 쉼표 제거 및 공백 제거
                G열값_정리 = G열값.replace(',', '').replace(' ', '')
                G열값 = float(G열값_정리) if G열값_정리 else 0
            else:
                G열값 = float(G열값) if G열값 else 0
        except (ValueError, TypeError):
            G열값 = 0
    
    # 정규화된 이름들로 키 생성 (주민번호와 함께)
    for 정규화된_이름 in 정규화된_이름들:
        키 = (정규화된_이름, 주민번호)
        
        # 데이터 누적
        사람별_데이터[키]['A열값들'].append(A열값)
        사람별_데이터[키]['G열합계'] += G열값
        사람별_데이터[키]['행데이터들'].append(행)
        사람별_데이터[키]['H열값들'].append(H열값)

# 병합된 데이터를 후처리 시트에 쓰기
새행번호 = 2  # 1행은 헤더

for 키, 데이터 in 사람별_데이터.items():
    이름, 주민번호 = 키
    A열값들 = 데이터['A열값들']
    G열합계 = 데이터['G열합계']
    행데이터들 = 데이터['행데이터들']
    H열값들 = 데이터['H열값들']
    
    # H열 값을 기준으로 정렬하여 가장 최신 데이터(가장 마지막 값) 선택
    # H열 값이 없는 경우 고려
    최신_행_인덱스 = 0
    if H열값들 and any(h for h in H열값들):
        # None 값 처리를 위해 빈 문자열로 대체
        정렬용_H열값들 = [(i, str(h) if h is not None else "") for i, h in enumerate(H열값들)]
        # 문자열 기준 내림차순 정렬하여 가장 최신 값 찾기
        정렬용_H열값들.sort(key=lambda x: x[1], reverse=True)
        최신_행_인덱스 = 정렬용_H열값들[0][0]
    
    # 가장 최신 행 데이터 사용
    최신_행 = 행데이터들[최신_행_인덱스]
    
    # A열에 중복값들을 쉼표로 구분하여 기록
    후처리시트.cell(row=새행번호, column=1, value=', '.join(str(값) for 값 in A열값들 if 값))
    
    # 나머지 열 복사 (B~F, H~) - 가장 최신 행 데이터 사용
    for i, 셀 in enumerate(최신_행):
        if i == 0:  # A열은 이미 처리함
            continue
        if i == 6:  # G열은 합계 값으로 대체
            후처리시트.cell(row=새행번호, column=7, value=G열합계)
        elif i == 7:  # H열은 '입금완료_' 텍스트 제거
            H열값 = 셀.value
            if H열값 and isinstance(H열값, str):
                H열값 = H열값.replace('입금완료_', '')
            후처리시트.cell(row=새행번호, column=i+1, value=H열값)
        else:
            후처리시트.cell(row=새행번호, column=i+1, value=셀.value)
    
    # 색상 변경 여부 확인
    색상변경 = None  # None: 색상 변경 없음, "오렌지": 오렌지색, "파랑": 파란색
    
    # F열이 빈 값이면 파란색으로 변경
    if 주민번호 is None or 주민번호 == '':
        색상변경 = "파랑"
    # F열 값이 13자리가 아닌 경우 오렌지색으로 변경
    elif isinstance(주민번호, str):
        if len(주민번호) != 13:
            색상변경 = "오렌지"
    else:
        # 숫자인 경우 문자열로 변환 후 확인
        if len(str(주민번호)) != 13:
            색상변경 = "오렌지"
    
    # 조건에 따라 색상 변경
    if 색상변경:
        for 열 in range(1, 원본시트.max_column + 1):
            if 색상변경 == "오렌지":
                후처리시트.cell(row=새행번호, column=열).fill = 오렌지색
            elif 색상변경 == "파랑":
                후처리시트.cell(row=새행번호, column=열).fill = 파란색
    
    새행번호 += 1

# 동명이인 찾기 및 저장
print("동명이인 찾는 중...")
이름별_주민번호들 = defaultdict(set)  # 이름별로 주민번호들을 저장

# 후처리 시트에서 이름별 주민번호 수집
for 행 in list(후처리시트.rows)[1:]:  # 헤더 제외
    if len(행) > 1 and len(행) > 5:  # B열과 F열이 있는지 확인
        이름 = 행[1].value if 행[1].value else ""
        주민번호 = 행[5].value if 행[5].value else ""
        
        if 이름 and 주민번호:
            이름별_주민번호들[이름].add(str(주민번호))

# 동명이인 찾기 (같은 이름에 다른 주민번호가 2개 이상인 경우)
동명이인_이름들 = {이름: 주민번호들 for 이름, 주민번호들 in 이름별_주민번호들.items() if len(주민번호들) > 1}

# 동명이인 데이터를 동명이인 시트에 저장
동명이인_행번호 = 2  # 1행은 헤더
for 이름, 주민번호들 in 동명이인_이름들.items():
    # 해당 이름의 모든 행을 후처리 시트에서 찾아서 동명이인 시트에 복사
    for 행 in list(후처리시트.rows)[1:]:  # 헤더 제외
        if len(행) > 1 and len(행) > 5:
            행_이름 = 행[1].value if 행[1].value else ""
            행_주민번호 = str(행[5].value) if 행[5].value else ""
            
            if 행_이름 == 이름 and 행_주민번호 in 주민번호들:
                # 행 데이터를 동명이인 시트에 복사
                for 열_인덱스, 셀 in enumerate(행):
                    동명이인시트.cell(row=동명이인_행번호, column=열_인덱스 + 1, value=셀.value)
                동명이인_행번호 += 1

# 열 너비 복사 (가능한 경우)
for 열 in range(1, 원본시트.max_column + 1):
    if 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width is not None:
        후처리시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width = 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width
        동명이인시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width = 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width

# 파일 저장
워크북.save(파일경로)
print(f"작업이 완료되었습니다. 동명이인 {len(동명이인_이름들)}명 발견: {list(동명이인_이름들.keys())}")
