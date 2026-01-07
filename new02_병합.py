#!/usr/bin/env python3
"""
가상환경 자동 감지 및 활성화
"""
import sys
import os
from pathlib import Path

# 현재 스크립트의 디렉토리
SCRIPT_DIR = Path(__file__).parent.absolute()
VENV_PYTHON = SCRIPT_DIR / 'venv' / 'bin' / 'python3'

# 가상환경이 존재하고 현재 Python이 가상환경이 아닌 경우
if VENV_PYTHON.exists() and 'venv' not in sys.executable:
    # 가상환경의 Python으로 재실행
    os.execv(str(VENV_PYTHON), [str(VENV_PYTHON)] + sys.argv)

import openpyxl
from openpyxl.styles import PatternFill
from collections import defaultdict
import os
import pandas as pd
from googleapiclient.discovery import build
from auth import get_credentials
from google_sheets_config import TRANSACTION_PREPROCESSING_FOLDER_ID

# 사용자로부터 년월 입력 받기 (예: 2512)
user_input = input("년월을 입력하세요 (예: 2512): ")

# 구글 인증 자격 증명 가져오기
print("구글 인증 중...")
creds = get_credentials()

# 구글 시트 API 및 드라이브 API 클라이언트 생성
sheets_service = build('sheets', 'v4', credentials=creds)
drive_service = build('drive', 'v3', credentials=creds)

# 구글 드라이브에서 년월 폴더 찾기
print(f"\n'{user_input}' 폴더 찾는 중...")
query = f"'{TRANSACTION_PREPROCESSING_FOLDER_ID}' in parents and mimeType='application/vnd.google-apps.folder' and name='{user_input}' and trashed=false"
existing_folders = drive_service.files().list(q=query, fields='files(id, name)').execute()

if not existing_folders.get('files'):
    print(f"❌ '{user_input}' 폴더를 찾을 수 없습니다.")
    exit()

date_folder_id = existing_folders['files'][0]['id']
print(f"✅ '{user_input}' 폴더를 찾았습니다.")

# 폴더 안에 스프레드시트 목록 가져오기
print(f"\n폴더 안에 스프레드시트 목록 가져오는 중...")
spreadsheet_query = f"'{date_folder_id}' in parents and mimeType='application/vnd.google-apps.spreadsheet' and trashed=false"
existing_spreadsheets = drive_service.files().list(q=spreadsheet_query, fields='files(id, name)').execute()

if not existing_spreadsheets.get('files'):
    print(f"❌ 폴더 안에 스프레드시트가 없습니다.")
    exit()

# 스프레드시트 목록 출력
print("\n사용 가능한 스프레드시트 목록:")
spreadsheet_list = existing_spreadsheets['files']
for i, spreadsheet in enumerate(spreadsheet_list, 1):
    print(f"{i}. {spreadsheet['name']}")

# 사용자 입력 받기
while True:
    try:
        선택 = int(input("\n처리할 스프레드시트 번호를 입력하세요: "))
        if 1 <= 선택 <= len(spreadsheet_list):
            target_spreadsheet = spreadsheet_list[선택-1]
            target_spreadsheet_id = target_spreadsheet['id']
            target_spreadsheet_name = target_spreadsheet['name']
            break
        else:
            print("올바른 번호를 입력해주세요.")
    except ValueError:
        print("숫자를 입력해주세요.")

print(f"\n선택된 스프레드시트: {target_spreadsheet_name}")

# 스프레드시트의 모든 시트 정보 가져오기
spreadsheet_info = sheets_service.spreadsheets().get(spreadsheetId=target_spreadsheet_id).execute()
sheets = spreadsheet_info.get('sheets', [])

# '구글시트'가 포함된 시트 찾기
원본시트_이름 = None
구글시트_시트들 = []
for sheet in sheets:
    sheet_title = sheet.get('properties', {}).get('title', '')
    if '구글시트' in sheet_title:
        구글시트_시트들.append(sheet_title)

if not 구글시트_시트들:
    print(f"❌ '구글시트'가 포함된 시트를 찾을 수 없습니다.")
    exit()

# '구글시트'가 포함된 시트가 여러 개인 경우 선택
if len(구글시트_시트들) > 1:
    print("\n'구글시트'가 포함된 시트 목록:")
    for i, sheet_name in enumerate(구글시트_시트들, 1):
        print(f"{i}. {sheet_name}")
    
    while True:
        try:
            선택 = int(input("\n처리할 시트 번호를 입력하세요: "))
            if 1 <= 선택 <= len(구글시트_시트들):
                원본시트_이름 = 구글시트_시트들[선택-1]
                break
            else:
                print("올바른 번호를 입력해주세요.")
        except ValueError:
            print("숫자를 입력해주세요.")
else:
    원본시트_이름 = 구글시트_시트들[0]
    print(f"✅ '{원본시트_이름}' 시트를 사용합니다.")

# 시트 데이터 가져오기
print(f"\n'{원본시트_이름}' 시트 데이터 읽는 중...")
result = sheets_service.spreadsheets().values().get(
    spreadsheetId=target_spreadsheet_id,
    range=f"{원본시트_이름}!A:Z"
).execute()

rows = result.get('values', [])
if not rows:
    print(f"❌ '{원본시트_이름}' 시트에 데이터가 없습니다.")
    exit()

# 데이터프레임으로 변환
df = pd.DataFrame(rows)
print(f"✅ {len(df)}행 데이터 읽기 완료")

# 임시 엑셀 파일로 저장하여 openpyxl로 처리
import tempfile
temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
temp_file_path = temp_file.name
temp_file.close()

df.to_excel(temp_file_path, sheet_name='25년3월데이터', index=False, header=False)

# 엑셀 파일 열기
워크북 = openpyxl.load_workbook(temp_file_path)

# 원본 시트 가져오기
원본시트 = 워크북['25년3월데이터']

# 새 시트 생성 (이미 있으면 삭제 후 생성)
if '후처리' in 워크북.sheetnames:
    워크북.remove(워크북['후처리'])
후처리시트 = 워크북.create_sheet('후처리')

# 동명이인 시트 생성 (이미 있으면 삭제 후 생성)
if '동명이인' in 워크북.sheetnames:
    워크북.remove(워크북['동명이인'])
동명이인시트 = 워크북.create_sheet('동명이인')

# 오렌지색 채우기 스타일 정의
오렌지색 = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
# 파란색 채우기 스타일 정의
파란색 = PatternFill(start_color="0000FF", end_color="0000FF", fill_type="solid")

# 헤더 추가 (명시적으로 설정)
헤더 = ['항목', '이름', '번호', '-', '계좌', '주민번호', '입금액', '상태']
for i, 헤더값 in enumerate(헤더, 1):
    후처리시트.cell(row=1, column=i, value=헤더값)
    동명이인시트.cell(row=1, column=i, value=헤더값)

# 동일인 데이터 그룹화 (이름과 주민번호가 같은 경우)
사람별_데이터 = defaultdict(lambda: {'A열값들': [], 'G열합계': 0, '행데이터들': [], 'H열값들': []})

# 이름에서 괄호 처리 함수
def 이름_정규화(이름):
    """이름에서 괄호를 제거하고 괄호 안의 이름도 고려하여 정규화"""
    if not 이름 or not isinstance(이름, str):
        return ""
    
    # 전체 이름에서 공백 제거
    이름 = 이름.replace(' ', '')
    
    # 괄호가 있는 경우
    if '(' in 이름 and ')' in 이름:
        # 괄호 밖 이름과 괄호 안 이름 모두 추출
        괄호밖_이름 = 이름.split('(')[0].strip()
        괄호안_이름 = 이름.split('(')[1].split(')')[0].strip()
        return [괄호밖_이름, 괄호안_이름]
    else:
        # 괄호가 없는 경우
        return [이름.strip()]

# 첫 번째 행은 이미 처리했으므로 두 번째 행부터 시작
for 행 in list(원본시트.rows)[1:]:
    # B열(이름)과 F열(주민번호) 값 가져오기
    원본_이름 = 행[1].value if len(행) > 1 else ""
    주민번호 = 행[5].value if len(행) > 5 else ""
    
    # 이름 정규화
    정규화된_이름들 = 이름_정규화(원본_이름)
    
    # A열 값과 G열 값, H열 값 가져오기
    A열값 = 행[0].value if len(행) > 0 else ""
    G열값 = 행[6].value if len(행) > 6 else 0
    H열값 = 행[7].value if len(행) > 7 else ""
    
    # 숫자가 아닌 경우 처리 - 쉼표 제거 및 숫자 변환 개선
    if not isinstance(G열값, (int, float)):
        try:
            if isinstance(G열값, str):
                # 쉼표 제거 및 공백 제거
                G열값_정리 = G열값.replace(',', '').replace(' ', '')
                G열값 = float(G열값_정리) if G열값_정리 else 0
            else:
                G열값 = float(G열값) if G열값 else 0
        except (ValueError, TypeError):
            G열값 = 0
    
    # 정규화된 이름들로 키 생성 (주민번호와 함께)
    for 정규화된_이름 in 정규화된_이름들:
        키 = (정규화된_이름, 주민번호)
        
        # 데이터 누적
        사람별_데이터[키]['A열값들'].append(A열값)
        사람별_데이터[키]['G열합계'] += G열값
        사람별_데이터[키]['행데이터들'].append(행)
        사람별_데이터[키]['H열값들'].append(H열값)

# 병합된 데이터를 후처리 시트에 쓰기
새행번호 = 2  # 1행은 헤더
색상정보 = []  # (행번호, 색상타입) 저장

for 키, 데이터 in 사람별_데이터.items():
    이름, 주민번호 = 키
    A열값들 = 데이터['A열값들']
    G열합계 = 데이터['G열합계']
    행데이터들 = 데이터['행데이터들']
    H열값들 = 데이터['H열값들']
    
    # H열 값을 기준으로 정렬하여 가장 최신 데이터(가장 마지막 값) 선택
    # H열 값이 없는 경우 고려
    최신_행_인덱스 = 0
    if H열값들 and any(h for h in H열값들):
        # None 값 처리를 위해 빈 문자열로 대체
        정렬용_H열값들 = [(i, str(h) if h is not None else "") for i, h in enumerate(H열값들)]
        # 문자열 기준 내림차순 정렬하여 가장 최신 값 찾기
        정렬용_H열값들.sort(key=lambda x: x[1], reverse=True)
        최신_행_인덱스 = 정렬용_H열값들[0][0]
    
    # 가장 최신 행 데이터 사용
    최신_행 = 행데이터들[최신_행_인덱스]
    
    # A열에 중복값들을 쉼표로 구분하여 기록
    후처리시트.cell(row=새행번호, column=1, value=', '.join(str(값) for 값 in A열값들 if 값))
    
    # 나머지 열 복사 (B~F, H~) - 가장 최신 행 데이터 사용
    for i, 셀 in enumerate(최신_행):
        if i == 0:  # A열은 이미 처리함
            continue
        if i == 6:  # G열은 합계 값으로 대체
            후처리시트.cell(row=새행번호, column=7, value=G열합계)
        elif i == 7:  # H열은 '입금완료_' 텍스트 제거
            H열값 = 셀.value
            if H열값 and isinstance(H열값, str):
                H열값 = H열값.replace('입금완료_', '')
            후처리시트.cell(row=새행번호, column=i+1, value=H열값)
        else:
            후처리시트.cell(row=새행번호, column=i+1, value=셀.value)
    
    # 색상 변경 여부 확인
    색상변경 = None  # None: 색상 변경 없음, "오렌지": 오렌지색, "파랑": 파란색
    
    # F열이 빈 값이면 파란색으로 변경
    if 주민번호 is None or 주민번호 == '':
        색상변경 = "파랑"
    # F열 값이 13자리가 아닌 경우 오렌지색으로 변경
    elif isinstance(주민번호, str):
        if len(주민번호) != 13:
            색상변경 = "오렌지"
    else:
        # 숫자인 경우 문자열로 변환 후 확인
        if len(str(주민번호)) != 13:
            색상변경 = "오렌지"
    
    # 조건에 따라 색상 변경 (엑셀 파일에 적용)
    if 색상변경:
        for 열 in range(1, 원본시트.max_column + 1):
            if 색상변경 == "오렌지":
                후처리시트.cell(row=새행번호, column=열).fill = 오렌지색
            elif 색상변경 == "파랑":
                후처리시트.cell(row=새행번호, column=열).fill = 파란색
        # 구글시트용 색상 정보 저장 (0-based 인덱스)
        색상정보.append((새행번호 - 1, 색상변경))  # 새행번호는 1-based, 구글시트는 0-based
    
    새행번호 += 1

# 동명이인 찾기 및 저장
print("동명이인 찾는 중...")
이름별_주민번호들 = defaultdict(set)  # 이름별로 주민번호들을 저장

# 후처리 시트에서 이름별 주민번호 수집
for 행 in list(후처리시트.rows)[1:]:  # 헤더 제외
    if len(행) > 1 and len(행) > 5:  # B열과 F열이 있는지 확인
        이름 = 행[1].value if 행[1].value else ""
        주민번호 = 행[5].value if 행[5].value else ""
        
        if 이름 and 주민번호:
            이름별_주민번호들[이름].add(str(주민번호))

# 동명이인 찾기 (같은 이름에 다른 주민번호가 2개 이상인 경우)
동명이인_이름들 = {이름: 주민번호들 for 이름, 주민번호들 in 이름별_주민번호들.items() if len(주민번호들) > 1}

# 동명이인 데이터를 동명이인 시트에 저장
동명이인_행번호 = 2  # 1행은 헤더
for 이름, 주민번호들 in 동명이인_이름들.items():
    # 해당 이름의 모든 행을 후처리 시트에서 찾아서 동명이인 시트에 복사
    for 행 in list(후처리시트.rows)[1:]:  # 헤더 제외
        if len(행) > 1 and len(행) > 5:
            행_이름 = 행[1].value if 행[1].value else ""
            행_주민번호 = str(행[5].value) if 행[5].value else ""
            
            if 행_이름 == 이름 and 행_주민번호 in 주민번호들:
                # 행 데이터를 동명이인 시트에 복사
                for 열_인덱스, 셀 in enumerate(행):
                    동명이인시트.cell(row=동명이인_행번호, column=열_인덱스 + 1, value=셀.value)
                동명이인_행번호 += 1

# 열 너비 복사 (가능한 경우)
for 열 in range(1, 원본시트.max_column + 1):
    if 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width is not None:
        후처리시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width = 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width
        동명이인시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width = 원본시트.column_dimensions[openpyxl.utils.get_column_letter(열)].width

# 임시 파일에 저장
워크북.save(temp_file_path)

# 결과를 구글시트에 업로드
print(f"\n구글시트에 결과 업로드 중...")

# 후처리 시트 데이터 읽기
후처리_df = pd.read_excel(temp_file_path, sheet_name='후처리')
동명이인_df = pd.read_excel(temp_file_path, sheet_name='동명이인')

# 기존 시트 목록 확인 (다시 가져오기)
spreadsheet_info = sheets_service.spreadsheets().get(spreadsheetId=target_spreadsheet_id).execute()
sheets = spreadsheet_info.get('sheets', [])
existing_sheet_names = {sheet.get('properties', {}).get('title', '') for sheet in sheets}

# 원본 시트명에서 타임스탬프 추출 (예: 전처리_구글시트_20260107_101058 -> 20260107_101058)
from datetime import datetime
import re

타임스탬프 = None
match = re.search(r'(\d{8}_\d{6})', 원본시트_이름)
if match:
    타임스탬프 = match.group(1)
else:
    # 타임스탬프가 없으면 현재 시간 사용
    타임스탬프 = datetime.now().strftime('%Y%m%d_%H%M%S')

후처리_시트명 = f'후처리_구글시트_{타임스탬프}'
동명이인_시트명 = f'동명이인_구글시트_{타임스탬프}'

# 시트명이 이미 존재하는지 확인하고 고유한 이름 생성
counter = 2
while 후처리_시트명 in existing_sheet_names:
    후처리_시트명 = f'후처리_구글시트_{타임스탬프}_{counter}'
    counter += 1

counter = 2
while 동명이인_시트명 in existing_sheet_names:
    동명이인_시트명 = f'동명이인_구글시트_{타임스탬프}_{counter}'
    counter += 1

# 새 시트 추가
add_sheet_request = {
    'requests': [
        {
            'addSheet': {
                'properties': {
                    'title': 후처리_시트명
                }
            }
        },
        {
            'addSheet': {
                'properties': {
                    'title': 동명이인_시트명
                }
            }
        }
    ]
}
sheets_service.spreadsheets().batchUpdate(
    spreadsheetId=target_spreadsheet_id,
    body=add_sheet_request
).execute()

# 후처리 시트 데이터 업로드 (헤더 포함)
헤더 = ['항목', '이름', '번호', '-', '계좌', '주민번호', '입금액', '상태']
후처리_values = 후처리_df.fillna('').astype(str).values.tolist()
# 헤더를 첫 번째 행으로 추가
후처리_values_with_header = [헤더] + 후처리_values
sheets_service.spreadsheets().values().update(
    spreadsheetId=target_spreadsheet_id,
    range=f'{후처리_시트명}!A1',
    valueInputOption='USER_ENTERED',
    body={'values': 후처리_values_with_header}
).execute()

# 후처리 시트 ID 가져오기
후처리_시트_id = None
spreadsheet_info = sheets_service.spreadsheets().get(spreadsheetId=target_spreadsheet_id).execute()
for sheet in spreadsheet_info.get('sheets', []):
    if sheet.get('properties', {}).get('title') == 후처리_시트명:
        후처리_시트_id = sheet.get('properties', {}).get('sheetId')
        break

# 색상 스타일 적용
if 후처리_시트_id and 색상정보:
    print(f"색상 스타일 적용 중...")
    format_requests = []
    
    for 행번호_0based, 색상타입 in 색상정보:
        # 행번호는 0-based 인덱스 (헤더는 0번째 행, 데이터는 1번째 행부터)
        if 색상타입 == "오렌지":
            # 오렌지색 배경 (RGB: 255, 165, 0)
            format_requests.append({
                'repeatCell': {
                    'range': {
                        'sheetId': 후처리_시트_id,
                        'startRowIndex': 행번호_0based,
                        'endRowIndex': 행번호_0based + 1,
                        'startColumnIndex': 0,
                        'endColumnIndex': len(헤더)  # 헤더 개수만큼
                    },
                    'cell': {
                        'userEnteredFormat': {
                            'backgroundColor': {
                                'red': 1.0,
                                'green': 0.647,
                                'blue': 0.0
                            }
                        }
                    },
                    'fields': 'userEnteredFormat.backgroundColor'
                }
            })
        elif 색상타입 == "파랑":
            # 파란색 배경 (RGB: 0, 0, 255)
            format_requests.append({
                'repeatCell': {
                    'range': {
                        'sheetId': 후처리_시트_id,
                        'startRowIndex': 행번호_0based,
                        'endRowIndex': 행번호_0based + 1,
                        'startColumnIndex': 0,
                        'endColumnIndex': len(헤더)  # 헤더 개수만큼
                    },
                    'cell': {
                        'userEnteredFormat': {
                            'backgroundColor': {
                                'red': 0.0,
                                'green': 0.0,
                                'blue': 1.0
                            }
                        }
                    },
                    'fields': 'userEnteredFormat.backgroundColor'
                }
            })
    
    if format_requests:
        sheets_service.spreadsheets().batchUpdate(
            spreadsheetId=target_spreadsheet_id,
            body={'requests': format_requests}
        ).execute()
        print(f"✅ 색상 스타일 적용 완료 ({len(색상정보)}개 행)")

# 동명이인 시트 데이터 업로드
동명이인_values = 동명이인_df.fillna('').astype(str).values.tolist()
if 동명이인_values:
    sheets_service.spreadsheets().values().update(
        spreadsheetId=target_spreadsheet_id,
        range=f'{동명이인_시트명}!A1',
        valueInputOption='USER_ENTERED',
        body={'values': 동명이인_values}
    ).execute()

# 임시 파일 삭제
os.unlink(temp_file_path)

spreadsheet_url = f"https://docs.google.com/spreadsheets/d/{target_spreadsheet_id}/edit"
print(f"✅ 작업이 완료되었습니다.")
print(f"동명이인 {len(동명이인_이름들)}명 발견: {list(동명이인_이름들.keys())}")
print(f"스프레드시트 URL: {spreadsheet_url}")
