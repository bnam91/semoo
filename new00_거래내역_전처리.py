from pathlib import Path
import pandas as pd
from openpyxl import load_workbook

# 현재 폴더 경로
current_dir = Path('.')

# '거래내역조회_'가 포함된 엑셀 파일 찾기 (현재 폴더에만)
excel_files = []
for file_path in current_dir.glob('*거래내역조회_*.xls*'):
    if file_path.is_file():  # 파일인지 확인
        excel_files.append(file_path)

# 파일명 기준으로 정렬
excel_files.sort()

# 번호를 붙여서 출력
print(f"총 {len(excel_files)}개의 파일을 찾았습니다:\n")
for idx, file_path in enumerate(excel_files, 1):
    print(f"{idx}. {file_path}")

# 사용자로부터 번호 입력 받기
if excel_files:
    try:
        choice = int(input("\n처리할 파일 번호를 선택하세요: "))
        if 1 <= choice <= len(excel_files):
            selected_file = excel_files[choice - 1]
            print(f"\n선택된 파일: {selected_file}")
            
            # 전처리 파일명 생성 (항상 .xlsx 확장자로 저장)
            original_path = Path(selected_file)
            copy_path = original_path.parent / f"{original_path.stem}_전처리.xlsx"
            
            # 엑셀 파일 읽기
            print("엑셀 파일 읽는 중...")
            df = pd.read_excel(selected_file, header=None)
            
            print(f"원본 파일 크기: {df.shape[0]}행 x {df.shape[1]}열")
            
            # 원본 파일의 D7:D999 합계 계산 (마지막 행은 SUM이므로 제외)
            original_d_sum = 0
            if df.shape[1] > 3 and len(df) > 6:  # D열이 있고 7행 이상인지 확인
                # D7부터 마지막 행-1까지 (마지막 행은 SUM이므로 제외)
                start_idx = 6  # D7 (인덱스 6)
                end_idx = len(df) - 2  # 마지막 행 제외 (인덱스는 len-2)
                if end_idx >= start_idx:
                    d_col_original = pd.to_numeric(df.iloc[start_idx:end_idx+1, 3], errors='coerce')
                    original_d_sum = d_col_original.sum()
                    print(f"\n원본 파일 D7:D{end_idx+1} 합계 (마지막 SUM 행 제외): {original_d_sum:,.0f}")
            
            # a. 1-5행 삭제 (인덱스 0-4)
            if len(df) >= 5:
                df = df.iloc[5:].reset_index(drop=True)
                print("1-5행 삭제 완료")
            
            # b. E, F, G열 삭제 (인덱스 4, 5, 6)
            columns_to_drop = []
            if df.shape[1] > 4:
                columns_to_drop.append(4)  # E열
            if df.shape[1] > 5:
                columns_to_drop.append(5)  # F열
            if df.shape[1] > 6:
                columns_to_drop.append(6)  # G열
            
            if columns_to_drop:
                df = df.drop(df.columns[columns_to_drop], axis=1)
                print(f"E, F, G열 삭제 완료")
            
            # c. 마지막 행 삭제
            if len(df) > 0:
                df = df.iloc[:-1].reset_index(drop=True)
                print("마지막 행 삭제 완료")
            
            print(f"처리 후 파일 크기: {df.shape[0]}행 x {df.shape[1]}열")
            
            # 제외할 키워드 리스트
            exclude_keywords = [
                # 개인명
                '신현빈', '김지수', '임채빈', '정철호', '전현선', '신운균', '전수현',
                # 회사/기관
                '회사', '아이플', '네이버', '디베스트컴퍼니', '애드온비', '피에스팀',
                # 보험/연금
                '산재보험', '국민연금', '국민건강',
                # 통신사
                'SKTL',
                # 카드사, 결제
                '삼성카드', '현대카드', '국민카드', 'GSPAY', 'GSPay', '페이',
                # 세금
                '지방세', '소득세', '부가가치세', '관세', '경찰청', '과태료',
                # 쇼핑몰
                '무신사', '쿠팡', '당근',
                # 기타
                '월세', '배송비', '29고6425', '대출', '대체', '지로',
                '고용보험', '카카오T', '이자'
            ]
            
            # 디버깅: 키워드 리스트 확인
            print(f"\n📋 제외 키워드 리스트 ({len(exclude_keywords)}개):")
            print(f"  - 전체 키워드: {exclude_keywords}")
            
            # C열(인덱스 2)에서 키워드가 포함된 행 찾기
            if df.shape[1] > 2:  # C열이 있는지 확인
                # C열의 값이 문자열인지 확인하고 키워드 포함 여부 체크 (대소문자 구분 없이)
                def check_keyword(cell_value):
                    cell_str = str(cell_value)
                    cell_str_upper = cell_str.upper()
                    for keyword in exclude_keywords:
                        keyword_upper = keyword.upper()
                        # 영문은 대소문자 구분 없이, 한글은 그대로 비교
                        if keyword in cell_str or keyword_upper in cell_str_upper:
                            return True
                    return False
                
                exclude_mask = df.iloc[:, 2].astype(str).apply(check_keyword)
                
                # 디버깅: 모든 키워드에 대한 매칭 테스트
                print(f"\n🔍 키워드 매칭 상세 테스트:")
                keyword_match_count = {keyword: 0 for keyword in exclude_keywords}
                
                for idx, row in df.iterrows():
                    if len(row) > 2:
                        c_value = str(row.iloc[2])
                        is_excluded = exclude_mask.iloc[idx] if idx < len(exclude_mask) else False
                        
                        # 각 키워드별로 매칭 확인
                        matched_keywords = []
                        for keyword in exclude_keywords:
                            if keyword in c_value or keyword.upper() in c_value.upper():
                                matched_keywords.append(keyword)
                                keyword_match_count[keyword] += 1
                        
                        if matched_keywords:
                            print(f"  - 행 {idx+1}: C열='{c_value}'")
                            print(f"    → 매칭된 키워드: {matched_keywords}")
                            print(f"    → exclude_mask 결과: {is_excluded}")
                            if not is_excluded:
                                print(f"    → ⚠️  키워드가 매칭되었지만 제외되지 않음!")
                
                # 키워드별 매칭 통계
                print(f"\n📊 키워드별 매칭 통계:")
                for keyword, count in keyword_match_count.items():
                    if count > 0:
                        print(f"  - '{keyword}': {count}개 행 매칭")
                
                # 제외할 행 (sheet2로 이동)
                df_excluded = df[exclude_mask].copy()
                # 남은 행 (sheet1에 유지)
                df_main = df[~exclude_mask].copy()
                
                print(f"\nC열 키워드 검사 결과:")
                print(f"- 제외할 행: {len(df_excluded)}개")
                print(f"- 유지할 행: {len(df_main)}개")
                
                # ExcelWriter를 사용하여 여러 시트에 저장
                with pd.ExcelWriter(copy_path, engine='openpyxl') as writer:
                    # 먼저 Sheet1과 Sheet2를 저장
                    df_main.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
                    df_excluded.to_excel(writer, sheet_name='Sheet2', index=False, header=False)
                
                # 파일을 다시 열어서 Sheet0에 수식 추가
                wb = load_workbook(copy_path)
                
                # Sheet0 생성
                if 'Sheet0' in wb.sheetnames:
                    ws0 = wb['Sheet0']
                else:
                    ws0 = wb.create_sheet('Sheet0', 0)  # 첫 번째 위치에 생성
                
                # 헤더 작성
                ws0['A1'] = '항목'
                ws0['B1'] = '금액'
                
                # 데이터 작성
                ws0['A2'] = 'Sheet1 D열 합계'
                ws0['A3'] = 'Sheet2 D열 합계'
                ws0['A4'] = '총합'
                
                # 수식 작성 (D열은 엑셀에서 4번째 열)
                # Sheet1의 D2부터 D999까지 합계
                ws0['B2'] = '=SUM(Sheet1!D2:D999)'
                # Sheet2의 D2부터 D999까지 합계
                ws0['B3'] = '=SUM(Sheet2!D2:D999)'
                # 총합
                ws0['B4'] = '=SUM(B2:B3)'
                
                # 저장
                wb.save(copy_path)
                wb.close()
                
                # 콘솔에 출력할 합계 계산 (참고용)
                sheet1_sum = 0
                sheet2_sum = 0
                
                if df_main.shape[1] > 3:  # D열이 있는지 확인
                    d_col_main = pd.to_numeric(df_main.iloc[:, 3], errors='coerce')
                    sheet1_sum = d_col_main.sum()
                
                if df_excluded.shape[1] > 3:  # D열이 있는지 확인
                    d_col_excluded = pd.to_numeric(df_excluded.iloc[:, 3], errors='coerce')
                    sheet2_sum = d_col_excluded.sum()
                
                total_sum = sheet1_sum + sheet2_sum
                
                print(f"\nD열 합계 (참고용):")
                print(f"- 원본 파일 D7:D999: {original_d_sum:,.0f}")
                print(f"- Sheet1: {sheet1_sum:,.0f}")
                print(f"- Sheet2: {sheet2_sum:,.0f}")
                print(f"- 총합: {total_sum:,.0f}")
                print("(Sheet0에는 수식으로 저장되었습니다)")
                
                print(f"\n처리 완료! 전처리 파일이 저장되었습니다: {copy_path}")
                print(f"- Sheet0: 합계 정보")
                print(f"- Sheet1: {len(df_main)}행 (일반 데이터)")
                print(f"- Sheet2: {len(df_excluded)}행 (제외된 데이터)")
            else:
                # C열이 없는 경우 기존 방식으로 저장
                df.to_excel(copy_path, index=False, header=False)
                print(f"\n처리 완료! 전처리 파일이 저장되었습니다: {copy_path}")
                print("(C열이 없어 키워드 검사를 수행하지 않았습니다)")
                print(f"\n원본 파일 D7:D999 합계: {original_d_sum:,.0f}")
        else:
            print("잘못된 번호입니다.")
    except ValueError:
        print("숫자를 입력해주세요.")
    except Exception as e:
        print(f"오류 발생: {e}")
else:
    print("처리할 파일이 없습니다.")

